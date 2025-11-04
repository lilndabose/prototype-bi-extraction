import pandas as pd
from countries import get_country_map, extract_country
from helper import get_filepath,get_real_filename,get_filepath_to_execute,update_file_status,create_extractions_table_if_not_exists,insert_extraction_data,create_extractions_questions_table_if_not_exists,insert_extraction_questions_data, calculate_station_scores, get_station_code_by_name,create_hse_variant_table_if_not_exists, insert_hse_variant_data
from fuzzywuzzy import fuzz
from openpyxl import load_workbook
from rapidfuzz import fuzz, process
from concurrent.futures import ThreadPoolExecutor, as_completed
import time

FILE_PATH = get_filepath_to_execute()
ORIGINAL_PATH = get_filepath("Invariants.xlsx")

start_time = time.time()

xls = pd.ExcelFile(FILE_PATH)
sheet_names = xls.sheet_names

def load_sheet(sheet_name):
    return sheet_name, pd.read_excel(xls, sheet_name=sheet_name)

sheets = {}
with ThreadPoolExecutor(max_workers=3) as executor:
    futures = [executor.submit(load_sheet, name) for name in ["Inspections", "Questions", "HSE Invariants"]]
    for future in as_completed(futures):
        sheet_name, df = future.result()
        sheets[sheet_name] = df

df = sheets.get("Inspections", pd.read_excel(xls, sheet_name=0))
questions_df = sheets["Questions"]
hse_variant_df = sheets["HSE Invariants"]

print(f"Chargee {len(sheets)} feuilles en {time.time() - start_time:.2f}s")

def clean_dataframe(df):
    """Nettoyer les colonnes et les valeurs d'un dataframe en une seule passe"""
    df.columns = df.columns.str.strip().str.lower()
    
    object_cols = df.select_dtypes(include="object").columns
    df[object_cols] = df[object_cols].apply(lambda x: x.str.strip() if x.dtype == "object" else x)
    
    return df

COUNTRY_MAP = get_country_map()

def process_affiliate_and_country(df):
    df["affiliate"] = df["affiliate"].apply(extract_country)
    df["country_code"] = df["affiliate"].str.lower().map(COUNTRY_MAP)
    return df

def prepare_data_for_db(df):
    """ nettoyer les donnees pour enlever les champs null """
    df_clean = df.copy()
    
    # Remplacer NaN avec None pour compatibiliter de type null dans la bd
    df_clean = df_clean.where(pd.notna(df_clean), None)
    
    for col in df_clean.select_dtypes(include=['float64', 'float32']).columns:
        if df_clean[col].notna().any():
            if (df_clean[col].dropna() % 1 == 0).all():
                df_clean[col] = df_clean[col].astype('Int64')
    
    return df_clean

print("Traitement des données d'extraction")
df = clean_dataframe(df)
filtered_df = df[df["inspector"].str.lower() == "all"].copy()
filtered_df = process_affiliate_and_country(filtered_df)

filtered_df = prepare_data_for_db(filtered_df)
extraction_data = filtered_df.to_dict(orient="records")

print("Traitement des données relatives aux questions")
questions_df = clean_dataframe(questions_df)
questions_df = process_affiliate_and_country(questions_df)

# Vectorized column filtering
columns_to_keep = ["zone", "sub-zone", "affiliate", "station name", "station code", "d.02", "ep11"]
questions_df = questions_df[[col for col in columns_to_keep if col in questions_df.columns]]

questions_df = prepare_data_for_db(questions_df)
questions_data = questions_df.to_dict(orient="records")

print("Calcul des scores des stations")
stations_scores = calculate_station_scores(questions_data)


print("Traitement des variantes HSE")
hse_variant_df = clean_dataframe(hse_variant_df)
hse_variant_df = process_affiliate_and_country(hse_variant_df)
hse_variant_dict = hse_variant_df.to_dict(orient="records")

# Create a lookup dictionary for HSE variant data by station code
hse_lookup = {}
for hse_row in hse_variant_dict:
    station_code = hse_row.get("station code")
    if station_code:
        hse_lookup[station_code] = hse_row

hse_variant_data = questions_df.to_dict(orient="records")

for row in hse_variant_data:
    station_code = row.get("station code")
    if station_code and station_code in stations_scores:
        row["ep11"] = stations_scores[station_code]
    else:
        row["ep11"] = None
        
    # Add all columns from hse_variant_df for this station
    if station_code and station_code in hse_lookup:
        hse_row_data = hse_lookup[station_code]
        # Add all EP, ES, ET, Score columns from hse_variant_df
        for key, value in hse_row_data.items():
            # Skip columns that are already in row (to avoid overwriting)
            if key not in row:
                row[key] = value

print("Création de la tables hse_variantes dans la base de données")
db_start = time.time()

def create_table_safe(table_func, data):
    try:
        if not data:
            print("Aucune donnée pour créer la table")
            return False
        
        # Filter out None keys and clean column names
        valid_keys = [k for k in data[0].keys() if k is not None and str(k).lower() != 'nan']
        table_func(valid_keys)
        return True
    except Exception as e:
        print(f"Erreur lors de la création de la table: {e}")
        return False

with ThreadPoolExecutor(max_workers=3) as executor:
    futures = [
        executor.submit(create_table_safe, create_extractions_table_if_not_exists, extraction_data),
        executor.submit(create_table_safe, create_extractions_questions_table_if_not_exists, questions_data),
        executor.submit(create_table_safe, create_hse_variant_table_if_not_exists, hse_variant_data)
    ]
    
    for future in as_completed(futures):
        future.result()

print(f"✅ Tables created in {time.time() - db_start:.2f}s")

print("💾 Inserting data into database...")
insert_start = time.time()

def clean_record_for_insertion(record):
    return {k: v for k, v in record.items() 
            if k is not None and str(k).lower() != 'nan' and pd.notna(k)}

def insert_with_logging(insert_func, data, name):
    try:
        # Pour les variantes HSE, aucun nettoyage supplémentaire n'est nécessaire 
        # car les données proviennent de questions_df, qui est déjà nettoyé.
        # Assurez-vous simplement qu'il n'y a pas de valeurs NaN dans les enregistrements.
        cleaned_data = []
        for record in data:
            cleaned_record = {}
            for k, v in record.items():
                # Skip None keys and 'nan' string keys
                if k is None or str(k).lower() == 'nan':
                    continue
                # Convert NaN values to None
                if isinstance(v, float) and pd.isna(v):
                    cleaned_record[k] = None
                else:
                    cleaned_record[k] = v
            if cleaned_record:
                cleaned_data.append(cleaned_record)
        
        if not cleaned_data:
            print(f"{name}: aucun records valide à insérer")
            return False
            
        insert_func(cleaned_data)
        print(f"  ✅ {name}: {len(cleaned_data)} records insérés")
        return True
    except Exception as e:
        print(f"{name} insertion echouer: {e}")
        if data:
            print(f"Sample record keys: {list(data[0].keys())[:10]}")
            print(f"Sample values: {list(data[0].values())[:5]}")
        return False

# Insérer toutes les données en parallèle
with ThreadPoolExecutor(max_workers=3) as executor:
    insert_futures = {
        executor.submit(insert_with_logging, insert_extraction_data, extraction_data, "Extraction data"): "extraction",
        executor.submit(insert_with_logging, insert_extraction_questions_data, questions_data, "Questions data"): "questions",
        executor.submit(insert_with_logging, insert_hse_variant_data, hse_variant_data, "HSE variant data"): "hse"
    }
    
    # Wait for all insertions to complete
    for future in as_completed(insert_futures):
        data_type = insert_futures[future]
        success = future.result()

print(f"Toutes les données insereer dans: {time.time() - insert_start:.2f}s")

print(f"\nToutes les données insereer dans: {time.time() - start_time:.2f}s")
print("=" * 50)

hse_update_data = calculate_station_scores(questions_data)
station_name_and_codes = get_station_code_by_name(hse_variant_data)

affiliate_station_map = {}
for station_name, station_code in station_name_and_codes.items():
    affiliate = station_code[:2].upper()
    if affiliate not in affiliate_station_map:
        affiliate_station_map[affiliate] = {}
    affiliate_station_map[affiliate][station_name.lower()] = station_code

wb = load_workbook(ORIGINAL_PATH)
ws = wb.active

header_row_idx = 5
affiliate_col_idx = None
cost_center_col_idx = None
name_col_idx = None

for col_idx, cell in enumerate(ws[header_row_idx], start=1):
    header_value = str(cell.value).lower().strip() if cell.value else ""
    
    if "affiliate" in header_value:
        affiliate_col_idx = col_idx
    elif "cost center" in header_value or "cost centre" in header_value:
        cost_center_col_idx = col_idx
    elif header_value == "name" or "name" in header_value:
        name_col_idx = col_idx

if not all([affiliate_col_idx, cost_center_col_idx, name_col_idx]):
    print(f"Colonne: Affiliate={affiliate_col_idx}, Cost Center={cost_center_col_idx}, Name={name_col_idx}")
    raise ValueError("Impossible de trouver les colonnes requises")

data_start_row = 6
rows_to_process = []

for row_idx in range(data_start_row, ws.max_row + 1):
    inv_station_name = str(ws.cell(row_idx, name_col_idx).value or "").strip()
    inv_affiliate = str(ws.cell(row_idx, affiliate_col_idx).value or "").strip()
    
    if inv_station_name and inv_affiliate:
        rows_to_process.append((row_idx, inv_station_name, inv_affiliate))

def match_station(row_data):
    row_idx, inv_station_name, inv_affiliate = row_data
    inv_affiliate_upper = inv_affiliate.upper()
    
    if inv_affiliate_upper not in affiliate_station_map:
        return None
    
    affiliate_stations = affiliate_station_map[inv_affiliate_upper]
    inv_name_lower = inv_station_name.lower()
    
    result = process.extractOne(
        inv_name_lower,
        affiliate_stations.keys(),
        scorer=fuzz.token_set_ratio,
        score_cutoff=70
    )
    
    if result:
        matched_name, score, _ = result
        station_code = affiliate_stations[matched_name]
        return (row_idx, station_code, inv_station_name, matched_name, score)
    
    return None

matches_found = 0
updates = []

max_workers = min(8, len(rows_to_process) // 10 + 1)

with ThreadPoolExecutor(max_workers=max_workers) as executor:
    future_to_row = {executor.submit(match_station, row_data): row_data for row_data in rows_to_process}
    
    for future in as_completed(future_to_row):
        result = future.result()
        if result:
            row_idx, station_code, orig_name, matched_name, score = result
            updates.append((row_idx, station_code))
            matches_found += 1
            
            if matches_found % 50 == 0:
                print(f"  ... {matches_found} résultats trouvés")

print(f"\nTotal des résultats trouvés: {matches_found}")
print(f"Écriture {len(updates)} mises à jour vers Excel")

for row_idx, station_code in updates:
    ws.cell(row_idx, cost_center_col_idx).value = station_code

wb.save(ORIGINAL_PATH)

update_file_status(get_real_filename(FILE_PATH),'completed')